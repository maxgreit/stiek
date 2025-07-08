from openpyxl import load_workbook
import pandas as pd
import logging
import os
from pathlib import Path
import re


class ExcelProcessor:
    """
    Een class voor het verwerken van Excel bestanden voor looncomponenten.
    
    Deze class biedt functionaliteit voor het opschonen, verwerken en beheren
    van Excel bestanden die looncomponent data bevatten.
    """
    
    def __init__(self, base_dir=None, relative_path=None):
        """
        Initialiseer de ExcelProcessor.
        
        Args:
            base_dir: Het basis directory pad. Als None, wordt het huidige werkdirectory gebruikt.
            relative_path: Het relatieve pad naar het Excel bestand vanaf base_dir. 
                          Als None, moet het volledige pad worden opgegeven bij get_df_from_excel.
        """
        self.base_dir = Path(base_dir) if base_dir else Path.cwd()
        self.logger = logging.getLogger(__name__)
        self.relative_path = relative_path
        self.default_excel_path = self.base_dir / relative_path if relative_path else None
    
    def clean_excel(self, file_path):
        """
        Schoon een Excel bestand op door onnodige rijen te verwijderen.
        
        Args:
            file_path: Het pad naar het Excel bestand
            
        Returns:
            str: Status van de opschoning ('cleaned' of 'already_cleaned')
            
        Raises:
            ValueError: Als kolomnamen niet gevonden kunnen worden
            Exception: Voor andere fouten tijdens het opschonen
        """
        file_path = Path(file_path)
        
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Controleer of het bestand al is opgeschoond
            if sheet.cell(row=1, column=1).value is not None:
                self.logger.info(f"Bestand is al opgeschoond: {file_path}")
                return "already_cleaned"

            # Vind de eerste rij met kolomnamen
            header_row = self._find_header_row(sheet)
            if header_row is None:
                raise ValueError("Kolomnamen niet gevonden in het Excel-bestand.")

            # Verwijder rijen boven de kolomnamen
            if header_row > 1:
                sheet.delete_rows(1, header_row - 1)

            # Verwijder filter indien aanwezig
            if sheet.auto_filter.ref:
                sheet.auto_filter.ref = None

            # Vind en verwijder lege rijen onderaan
            last_data_row = self._find_last_data_row(sheet, header_row)
            if last_data_row < sheet.max_row:
                sheet.delete_rows(last_data_row + 1, sheet.max_row - last_data_row)

            # Sla het bestand op
            workbook.save(file_path)
            self.logger.info(f"Bestand opgeschoond en opgeslagen: {file_path}")
            return "cleaned"
            
        except Exception as e:
            self.logger.error(f"Fout tijdens het opschonen van het Excel-bestand: {e}")
            raise
    
    def _find_header_row(self, sheet):
        """
        Vind de eerste rij met kolomnamen.
        
        Args:
            sheet: Het Excel werkblad
            
        Returns:
            int: Het rijnummer van de header, of None als niet gevonden
        """
        for i, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), start=1):
            if row[0].value is not None:
                return i
        return None
    
    def _find_last_data_row(self, sheet, header_row):
        """
        Vind de laatste rij met data.
        
        Args:
            sheet: Het Excel werkblad
            header_row: Het rijnummer van de header
            
        Returns:
            int: Het rijnummer van de laatste rij met data
        """
        last_data_row = header_row
        for i, row in enumerate(sheet.iter_rows(min_row=header_row + 1, min_col=1, max_col=1), start=header_row + 1):
            if row[0].value is None:
                last_data_row = i - 1
                break
        return last_data_row

    def process_excel_file(self, filepath):
        """
        Verwerk een Excel bestand naar een pandas DataFrame.
        
        Args:
            filepath: Het pad naar het Excel bestand
            
        Returns:
            pandas.DataFrame: Het verwerkte DataFrame, of None bij fout
        """
        filepath = Path(filepath)
        
        try:
            self.logger.info("Start excel verwerking")
            df = pd.read_excel(filepath)
            self.logger.info("Excel bestand succesvol verwerkt")
            return df
        except Exception as e:
            self.logger.error(f"Fout bij het verwerken van het Excel bestand: {e}")
            return None

    def delete_excel_file(self, file_path):
        """
        Verwijder een Excel bestand.
        
        Args:
            file_path: Het pad naar het te verwijderen bestand
            
        Returns:
            bool: True als succesvol verwijderd, False anders
        """
        file_path = Path(file_path)
        
        try:
            if file_path.exists():
                file_path.unlink()
                self.logger.info(f"Excel bestand verwijderd: {file_path}")
                return True
            else:
                self.logger.info(f"Het bestand {file_path} bestaat niet of is al verwijderd.")
                return False
        except Exception as e:
            self.logger.error(f"Fout bij het verwijderen van het bestand: {e}")
            return False

    def find_file_by_pattern(self, directory, pattern):
        """
        Vind een bestand in een directory op basis van een regex patroon en extraheer groepen.
        
        Args:
            directory (str or Path): De directory om in te zoeken.
            pattern (str): Het regex patroon voor de bestandsnaam.
            
        Returns:
            Tuple[Path, tuple]: Het pad naar het gevonden bestand en een tuple met de 
                                 geëxtraheerde groepen uit het patroon.
                                 Retourneert (None, None) als er geen bestand wordt gevonden.
        """
        directory = Path(directory)
        if not directory.exists():
            self.logger.error(f"De directory '{directory}' bestaat niet.")
            return None, None
        
        # Gebruik een gecompileerd regex object voor efficiëntie
        regex = re.compile(pattern)
        
        for item in directory.iterdir():
            if item.is_file():
                match = regex.match(item.name)
                if match:
                    self.logger.info(f"Bestand gevonden op basis van patroon: {item}")
                    # Retourneer het pad en de gevonden groepen
                    return item, match.groups()
        
        self.logger.warning(f"Geen bestand gevonden in '{directory}' dat voldoet aan het patroon '{pattern}'.")
        return None, None

    def get_df_from_excel(self, custom_filepath=None):
        """
        Hoofdmethode voor het ophalen en verwerken van Excel data.
        
        Args:
            custom_filepath: Optioneel aangepast bestandspad. 
                           Als None, wordt het standaard pad gebruikt (indien geconfigureerd).
            
        Returns:
            Tuple[pandas.DataFrame, Path]: Het DataFrame en bestandspad, 
                                         of None bij fout
        """
        if custom_filepath is None and self.default_excel_path is None:
            self.logger.error("Geen bestandspad opgegeven en geen standaard pad geconfigureerd")
            return None
            
        filepath = Path(custom_filepath) if custom_filepath else self.default_excel_path
        
        if not filepath.exists():
            self.logger.error(f"Excel bestand niet gevonden: {filepath}")
            return None

        # Excel bestand opschonen
        try:
            status = self.clean_excel(filepath)
            if status == "already_cleaned":
                self.logger.info("Het bestand was al opgeschoond, doorgaan met de rest van het script.")
        except Exception as e:
            self.logger.error(f"Excel bestand opschonen mislukt: {e}")
            return None

        # Excel verwerking
        try:
            df = self.process_excel_file(filepath)
        except Exception as e:
            self.logger.error(f"Excel verwerking mislukt: {e}")
            return None

        # DataFrame validatie
        if df is None:
            self.logger.error("Geen DataFrame geretourneerd")
            return None
            
        if df.empty:
            self.logger.error("DataFrame is leeg")
            return None

        return df, filepath