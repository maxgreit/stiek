from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import logging
import os
import re

def get_file_path(base_dir):
    directory = os.path.join(base_dir, "e-uur/plaatsing/plaatsing_file")
    
    # Controleer of de directory bestaat
    if not os.path.exists(directory):
        raise FileNotFoundError(f"De directory '{directory}' bestaat niet.")
    
    # Controleer of de directory geen lege map is
    if not os.listdir(directory):
        raise FileNotFoundError(f"De directory '{directory}' is leeg. Er zijn geen bestanden om te verwerken.")
    
    filepath, detail, begindatum, einddatum = None, None, None, None

    for bestand in os.listdir(directory):
        # Controleer of het bestand voldoet aan het patroon "data_begindatum_einddatum.xlsx"
        if re.match(r"_\d{4}-\d{2}-\d{2}_\d{4}-\d{2}-\d{2}\.xlsx", bestand):
            filepath = os.path.join(directory, bestand)
            print(f"Gevonden bestand: {filepath}")
            
            # Extract begindatum en einddatum uit de bestandsnaam
            match = re.search(r"data_(\d{4}-\d{2}-\d{2})_(\d{4}-\d{2}-\d{2})\.xlsx", bestand)
            if match:
                begindatum = datetime.strptime(match.group(1), "%Y-%m-%d").date()
                einddatum = datetime.strptime(match.group(2), "%Y-%m-%d").date()
                
                detail = f"Data van {begindatum} tot {einddatum}"
                break  # Stop na het vinden van het eerste bestand dat voldoet
    
    # Controleer of er een geschikt bestand is gevonden
    if filepath is None:
        raise FileNotFoundError("Geen bestand gevonden dat voldoet aan het patroon 'data_begindatum_einddatum.xlsx' in de directory.")
    
    return filepath, detail, begindatum, einddatum

def clean_excel(file_path):
    try:
        # Laad het werkblad
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Controleer of de eerste cel in kolom A al gevuld is, zo ja, sla opschonen over
        if sheet.cell(row=1, column=1).value is not None:
            print(f"Bestand is al opgeschoond: {file_path}")
            return "already_cleaned"

        # Zoek de eerste rij met kolomnamen op basis van de eerste niet-lege cel in kolom A
        header_row = None
        for i, row in enumerate(sheet.iter_rows(min_col=1, max_col=1), start=1):
            if row[0].value is not None:
                header_row = i
                break

        if header_row is None:
            raise ValueError("Kolomnamen niet gevonden in het Excel-bestand.")

        # Verwijder de regels boven de kolomnamen
        if header_row > 1:
            sheet.delete_rows(1, header_row - 1)

        # Verwijder de filter als deze is toegepast
        if sheet.auto_filter.ref:
            sheet.auto_filter.ref = None

        # Identificeer de laatste rij met uren op basis van de eerste lege cel in kolom A onder de data
        last_data_row = header_row
        for i, row in enumerate(sheet.iter_rows(min_row=header_row + 1, min_col=1, max_col=1), start=header_row + 1):
            if row[0].value is None:
                last_data_row = i - 1
                break

        # Verwijder alle rijen onder de laatste rij met data
        sheet.delete_rows(last_data_row + 1, sheet.max_row - last_data_row)

        # Opslaan van het bestand
        workbook.save(file_path)
        print(f"Bestand opgeschoond en opgeslagen: {file_path}")
    except Exception as e:
        print(f"FOUTMELDING | Fout tijdens het opschonen van het Excel-bestand: {e}")
        raise

def process_excel_file(filepath):
    # log
    logging.info("Start excel verwerking")

    # Excel bestand verwerken
    try:
        df = pd.read_excel(filepath)
        logging.info("Excel bestand succesvol verwerkt")
        return df
    except Exception as e:
        logging.error(f"Fout bij het verwerken van het Excel bestand: {e}")
        return None
    
def delete_excel_file(file_path):
    try:
        # Controleer of het bestand bestaat
        if os.path.exists(file_path):
            os.remove(file_path)
            logging.info("Excel bestand verwijderd")
        else:
            logging.warning(f"Het bestand {file_path} bestaat niet of is al verwijderd.")
    except Exception as e:
        logging.error(f"Fout bij het verwijderen van het bestand: {e}")
        
def get_df_from_excel(base_dir):
    # Excel bestand ophalen
    bron = 'E-Uur'
    filepath = os.path.join(base_dir, "e-uur/plaatsing/plaatsing_file/Plaatsing.xlsx")

    # Excel bestand opschonen
    try:
        status = clean_excel(filepath)
        if status == "already_cleaned":
            logging.warning("Het bestand was al opgeschoond, doorgaan met de rest van het script.")
    except Exception as e: 
        logging.error(f"Excel bestand opschonen mislukt: {e}")
        return

    # Voer Excel transformatie uit
    try:
        df = process_excel_file(filepath)
    except Exception as e:
        logging.error(f"Excel verwerking mislukt: {e}")
        
    # Dataframe check
    if df is None:
        logging.error(f"Geen DataFrame geretourneerd")
        return
    if df.empty:
        logging.error(f"DataFrame is leeg")
        return
        
    return df, filepath